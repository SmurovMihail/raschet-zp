# (интерфейс)

import customtkinter as ct
from PIL import Image
import os
import sys
import tksheet
from tkinter import messagebox
from datetime import datetime

from database import Database
from payroll import PayrollCalculator

ct.set_appearance_mode("light")
ct.set_default_color_theme("blue")

class Main(ct.CTk):

    def __init__(self):
        super().__init__()

        self.fonts = {
            "header": ct.CTkFont(family="Arial", size=50, weight="bold"),
            "normal": ct.CTkFont(family="Arial", size=40),
            "button": ct.CTkFont(family="Arial", size=30, weight="bold"),
            "table": ("Arial", 20, "normal"),
            "small": ct.CTkFont(family="Arial", size=20)
        }

        self.title("Расчет заработной платы")
        
        self.db = Database()
        self.calc = PayrollCalculator(self.db)
        
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        self.geometry(f"{screen_width}x{screen_height}+0+0")

        # левое меню
        self.menu_frame = ct.CTkFrame(self)
        self.menu_frame.pack(side='left', fill='y')
        
        self.icon_image = self.load_icon("icon-homepage.png", 100)
        self.icon_label = ct.CTkLabel(self.menu_frame, text="", image=self.icon_image)
        self.icon_label.pack(padx=5, pady=5)

        self.staff_image = self.load_icon("icon-staff.png", 70)
        self.staff_b = ct.CTkButton(self.menu_frame, text="", image=self.staff_image, 
                                    command=lambda: self.switch_tab('Сотрудники'),
                                    corner_radius=20, width=70, height=100)
        self.staff_b.pack(padx=5, pady=5)

        self.calculation_image = self.load_icon("icon-calculation.png", 70)
        self.calculation_b = ct.CTkButton(self.menu_frame, text="", image=self.calculation_image, 
                                          command=lambda: self.switch_tab('Зарплата'), 
                                          corner_radius=20, width=70, height=100)
        self.calculation_b.pack(padx=5, pady=5)

        self.vedomosti_image = self.load_icon("icon-vedomosti.png", 70)
        self.vedomosti_b = ct.CTkButton(self.menu_frame, text="", image=self.vedomosti_image, 
                                        command=lambda: self.switch_tab('Ведомость'), 
                                        corner_radius=20, width=70, height=100)
        self.vedomosti_b.pack(padx=5, pady=5)

        self.report_image = self.load_icon("icon-report.png", 70)
        self.report_b = ct.CTkButton(self.menu_frame, text="", image=self.report_image, 
                                     command=lambda: self.switch_tab('Отчет'), 
                                     corner_radius=20, width=70, height=100)
        self.report_b.pack(padx=5, pady=5)

        # верхняя панель
        self.title_frame = ct.CTkFrame(self)
        self.title_frame.pack(side='top', fill='x', padx=10)

        self.title_lab = ct.CTkLabel(self.title_frame, text='Добро пожаловать', font=self.fonts['header'])
        self.title_lab.pack(side='left', padx=10, pady=5)

        # вкладки
        self.body_tabview = ct.CTkTabview(self) 
        self.body_tabview.pack(side="top", fill="both", expand=True, padx=10)

        for name in ["Сотрудники", "Зарплата", "Ведомость", "Отчет"]:
            self.body_tabview.add(name)

        self.body_tabview._segmented_button.grid_forget()

        self.setup_staff_tab()
        self.setup_payroll_tab()
        self.setup_vedomosti_tab()
        self.setup_report_tab()

        self.load_employees_data()
        self.load_calculation_data()
        self.load_vedomosti_data()
        self.load_report_data()

    def switch_tab(self, tab_name):
        """переключение между вкладками"""
        self.body_tabview.set(tab_name)

    # --- вкладка сотрудники ---
    def setup_staff_tab(self):
        staff_tab = self.body_tabview.tab("Сотрудники")
        
        header_frame = ct.CTkFrame(staff_tab, height=50, corner_radius=0)
        header_frame.pack(fill='x')

        ct.CTkLabel(header_frame, text='Сотрудники', font=self.fonts['normal']).pack(side='left', padx=5, pady=5)

        ct.CTkButton(header_frame, text="Экспорт", height=35, corner_radius=15, 
                    font=self.fonts['button'], command=self.export_employees).pack(side="right")
        
        ct.CTkButton(header_frame, text="Восстановить", height=35, corner_radius=15,
                    font=self.fonts['button'], fg_color="#2e7d32", hover_color="#1b5e20",
                    command=self.reinstate_employee).pack(side="right", padx=5)
        
        ct.CTkButton(header_frame, text="Уволить", height=35, corner_radius=15,
                    font=self.fonts['button'], fg_color="#c62828", hover_color="#b71c1c",
                    command=self.fire_employee).pack(side="right", padx=5)
        
        ct.CTkButton(header_frame, text="Добавить сотрудника", height=35, corner_radius=15,
                    font=self.fonts['button'], command=self.open_add_employee_dialog).pack(side="right", padx=10)

        self.staff_sheet = tksheet.Sheet(staff_tab, data=[])
        self.staff_sheet.pack(fill="both", expand=True)
        self.staff_sheet.set_options(font=self.fonts['table'])
        self.staff_sheet.enable_bindings("single_select", "copy", "cut", "paste")

    def load_employees_data(self):
        """загрузка всех сотрудников (включая уволенных)"""
        employees = self.db.get_all_employees(active_only=False)
        data = [["ID", "ФИО", "ИНН", "СНИЛС", "Должность", "Тип оплаты", "Оклад/Ставка", "Вычет", "Дата приема", "Статус"]]
        
        for emp in employees:
            if emp['тип_оплаты'] == 'оклад':
                rate_value = f"{emp['оклад']:,.0f} ₽"
                pay_type = "Оклад"
            else:
                rate_value = f"{emp['ставка_час']:,.0f} ₽/час"
                pay_type = "Почасовая"
            
            status = "Уволен" if emp.get('is_active', 1) == 0 else "Активен"
            
            data.append([
                emp['id'], emp['ФИО'], emp['ИНН'] or "-", emp['СНИЛС'] or "-", emp['должность'],
                pay_type, rate_value, f"{emp['стандартный_вычет']:,.0f} ₽", emp['дата_приёма'], status
            ])
        
        self.staff_sheet.set_sheet_data(data)
        self.staff_sheet.set_all_column_widths()

    def fire_employee(self):
        """увольнение выбранного сотрудника"""
        selected = self.staff_sheet.get_currently_selected()
        if not selected:
            messagebox.showerror("Ошибка", "Выберите сотрудника для увольнения")
            return
        
        sheet_data = self.staff_sheet.get_sheet_data()
        
        if isinstance(selected, tuple):
            row_idx = selected[0]
        else:
            row_idx = selected
        
        if row_idx == 0:
            messagebox.showerror("Ошибка", "Выберите строку с сотрудником")
            return
        
        if row_idx >= len(sheet_data):
            return
        
        employee_id = sheet_data[row_idx][0]
        employee_name = sheet_data[row_idx][1]
        
        current_status = sheet_data[row_idx][9] if len(sheet_data[row_idx]) > 9 else "Активен"
        if current_status == "Уволен":
            messagebox.showwarning("Предупреждение", f"Сотрудник {employee_name} уже уволен")
            return
        
        if messagebox.askyesno("Подтверждение", f"Уволить сотрудника {employee_name}?"):
            try:
                self.db.fire_employee(employee_id)
                messagebox.showinfo("Успех", f"Сотрудник {employee_name} уволен")
                self.load_employees_data()
                self.load_calculation_data()
                self.load_vedomosti_data()
                self.load_report_data()
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))

    def reinstate_employee(self):
        """восстановление уволенного сотрудника"""
        selected = self.staff_sheet.get_currently_selected()
        if not selected:
            messagebox.showerror("Ошибка", "Выберите сотрудника для восстановления")
            return
        
        sheet_data = self.staff_sheet.get_sheet_data()
        
        if isinstance(selected, tuple):
            row_idx = selected[0]
        else:
            row_idx = selected
        
        if row_idx == 0:
            messagebox.showerror("Ошибка", "Выберите строку с сотрудником")
            return
        
        if row_idx >= len(sheet_data):
            return
        
        employee_id = sheet_data[row_idx][0]
        employee_name = sheet_data[row_idx][1]
        
        current_status = sheet_data[row_idx][9] if len(sheet_data[row_idx]) > 9 else "Активен"
        if current_status == "Активен":
            messagebox.showwarning("Предупреждение", f"Сотрудник {employee_name} уже активен")
            return
        
        if messagebox.askyesno("Подтверждение", f"Восстановить сотрудника {employee_name}?"):
            try:
                self.db.reinstate_employee(employee_id)
                messagebox.showinfo("Успех", f"Сотрудник {employee_name} восстановлен")
                self.load_employees_data()
                self.load_calculation_data()
                self.load_vedomosti_data()
                self.load_report_data()
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))

    def open_add_employee_dialog(self):
        dialog = ct.CTkToplevel(self)
        dialog.title("Добавление сотрудника")
        dialog.geometry("750x550")
        dialog.grab_set()
        
        main_frame = ct.CTkFrame(dialog)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ct.CTkLabel(main_frame, text="Данные нового сотрудника", font=("Arial", 20, "bold")).pack(pady=(0, 15))
        
        def validate_digits(char):
            return char.isdigit() or char == ''
        
        validate_cmd = dialog.register(validate_digits)
        
        form_frame = ct.CTkFrame(main_frame)
        form_frame.pack(fill="both", expand=True)
        
        left_col = ct.CTkFrame(form_frame)
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 10))
        right_col = ct.CTkFrame(form_frame)
        right_col.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        # левая колонка
        ct.CTkLabel(left_col, text="ФИО *:", font=self.fonts['small'], anchor="w").pack(fill="x", pady=(0, 2))
        name_entry = ct.CTkEntry(left_col, font=self.fonts['small'])
        name_entry.pack(fill="x", pady=(0, 12))
        
        ct.CTkLabel(left_col, text="ИНН (10 или 12 цифр):", font=self.fonts['small'], anchor="w").pack(fill="x", pady=(0, 2))
        inn_entry = ct.CTkEntry(left_col, font=self.fonts['small'])
        inn_entry.configure(validate="key", validatecommand=(validate_cmd, '%S'))
        inn_entry.pack(fill="x", pady=(0, 12))
        
        ct.CTkLabel(left_col, text="СНИЛС (11 цифр):", font=self.fonts['small'], anchor="w").pack(fill="x", pady=(0, 2))
        snils_entry = ct.CTkEntry(left_col, font=self.fonts['small'])
        snils_entry.configure(validate="key", validatecommand=(validate_cmd, '%S'))
        snils_entry.pack(fill="x", pady=(0, 12))
        
        ct.CTkLabel(left_col, text="Должность *:", font=self.fonts['small'], anchor="w").pack(fill="x", pady=(0, 2))
        position_entry = ct.CTkEntry(left_col, font=self.fonts['small'])
        position_entry.pack(fill="x", pady=(0, 12))
        
        ct.CTkLabel(left_col, text="Дата приема *:", font=self.fonts['small'], anchor="w").pack(fill="x", pady=(0, 2))
        date_entry = ct.CTkEntry(left_col, font=self.fonts['small'])
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        date_entry.pack(fill="x", pady=(0, 12))
        
        # правая колонка
        ct.CTkLabel(right_col, text="Тип оплаты:", font=self.fonts['small'], anchor="w").pack(fill="x", pady=(0, 2))
        payment_type = ct.CTkComboBox(right_col, values=["оклад", "почасовая"], font=self.fonts['small'])
        payment_type.set("оклад")
        payment_type.pack(fill="x", pady=(0, 12))
        
        dynamic_frame = ct.CTkFrame(right_col)
        dynamic_frame.pack(fill="x", pady=(0, 12))
        
        salary_entry = ct.CTkEntry(dynamic_frame, font=self.fonts['small'], placeholder_text="Оклад (руб)")
        salary_entry.pack(fill="x", pady=(0, 5))
        salary_entry.insert(0, "0")
        
        rate_entry = ct.CTkEntry(dynamic_frame, font=self.fonts['small'], placeholder_text="Часовая ставка (руб)")
        rate_entry.pack(fill="x")
        rate_entry.insert(0, "0")
        rate_entry.pack_forget()
        
        ct.CTkLabel(right_col, text="Стандартный вычет (руб):", font=self.fonts['small'], anchor="w").pack(fill="x", pady=(0, 2))
        deduction_entry = ct.CTkEntry(right_col, font=self.fonts['small'])
        deduction_entry.insert(0, "0")
        deduction_entry.pack(fill="x", pady=(0, 12))
        
        def on_payment_type_change(choice):
            if choice == "оклад":
                salary_entry.pack(fill="x", pady=(0, 5))
                rate_entry.pack_forget()
            else:
                rate_entry.pack(fill="x", pady=(0, 5))
                salary_entry.pack_forget()
        
        payment_type.configure(command=on_payment_type_change)
        
        def save():
            inn = inn_entry.get().strip()
            snils = snils_entry.get().strip()
            
            if inn and inn not in ['', '-']:
                if len(inn) not in [10, 12] or not inn.isdigit():
                    messagebox.showerror("Ошибка", "ИНН должен содержать 10 или 12 цифр")
                    return
            
            if snils and snils not in ['', '-']:
                if len(snils) != 11 or not snils.isdigit():
                    messagebox.showerror("Ошибка", "СНИЛС должен содержать 11 цифр")
                    return
            
            data = {
                'ФИО': name_entry.get(),
                'ИНН': inn,
                'СНИЛС': snils,
                'должность': position_entry.get(),
                'тип_оплаты': payment_type.get(),
                'оклад': float(salary_entry.get() or 0),
                'ставка_час': float(rate_entry.get() or 0),
                'стандартный_вычет': float(deduction_entry.get() or 0),
                'дата_приёма': date_entry.get()
            }
            if not data['ФИО'] or not data['должность']:
                messagebox.showerror("Ошибка", "Заполните ФИО и Должность")
                return
            try:
                self.db.add_employee(data)
                messagebox.showinfo("Успех", "Сотрудник добавлен")
                dialog.destroy()
                self.load_employees_data()
                self.load_calculation_data()
                self.load_vedomosti_data()
                self.load_report_data()
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))
        
        btn_frame = ct.CTkFrame(main_frame)
        btn_frame.pack(fill="x", pady=(20, 0))

        ct.CTkButton(btn_frame, text="Сохранить", command=save, 
                height=40, corner_radius=10, font=self.fonts['button'],
                fg_color="#2e7d32", hover_color="#1b5e20").pack(side="left", padx=5, expand=True)

        ct.CTkButton(btn_frame, text="Отмена", command=dialog.destroy,
                height=40, corner_radius=10, font=self.fonts['button'],
                fg_color="#c62828", hover_color="#b71c1c").pack(side="left", padx=5, expand=True)

    # --- вкладка зарплата ---
    def setup_payroll_tab(self):
        payroll_tab = self.body_tabview.tab("Зарплата")
        
        header_frame = ct.CTkFrame(payroll_tab, height=50, corner_radius=0)
        header_frame.pack(fill='x')
        
        ct.CTkLabel(header_frame, text='Расчет зарплаты', font=self.fonts['normal']).pack(side='left', padx=5, pady=5)
        
        # выбор периода
        period_frame = ct.CTkFrame(header_frame)
        period_frame.pack(side="right", padx=10)
        
        ct.CTkLabel(period_frame, text="Год:", font=self.fonts['small']).pack(side="left", padx=5)
        self.year_var = ct.StringVar(value=str(datetime.now().year))
        self.year_combo = ct.CTkComboBox(period_frame, values=[str(y) for y in range(2023, 2027)],
                                          variable=self.year_var, width=80, command=self.on_year_change)
        self.year_combo.pack(side="left", padx=5)
        
        ct.CTkLabel(period_frame, text="Месяц:", font=self.fonts['small']).pack(side="left", padx=5)
        self.month_var = ct.StringVar(value=datetime.now().strftime("%m"))
        months = [f"{i:02d}" for i in range(1, 13)]
        self.month_combo = ct.CTkComboBox(period_frame, values=months,
                                           variable=self.month_var, width=80, command=self.on_month_change)
        self.month_combo.pack(side="left", padx=5)
        
        ct.CTkButton(header_frame, text="Рассчитать", height=35, corner_radius=15,
                    font=self.fonts['button'], command=self.calculate_payroll).pack(side="right", padx=10)
        ct.CTkButton(header_frame, text="Экспорт", height=35, corner_radius=15,
                    font=self.fonts['button'], command=self.export_calculation).pack(side="right")
        
        # таблица
        self.calc_sheet = tksheet.Sheet(payroll_tab, data=[])
        self.calc_sheet.pack(fill="both", side='left', expand=True)
        self.calc_sheet.set_options(font=self.fonts['table'])
        self.calc_sheet.enable_bindings("single_select", "copy", "cut", "paste", "edit_cell")
        self.calc_sheet.extra_bindings([("cell_edited", self.on_cell_edited)])
        
        # панель итогов
        info_frame = ct.CTkFrame(payroll_tab, width=220, corner_radius=0)
        info_frame.pack(fill='both', side="right", padx=10)
        
        frame1 = ct.CTkFrame(info_frame, corner_radius=10)
        frame1.pack(fill='both', padx=10, pady=5)
        ct.CTkLabel(frame1, text='Начислено всего:', font=self.fonts['small']).pack(padx=5, pady=5)
        self.total_accrued_label = ct.CTkLabel(frame1, text='0 ₽', font=self.fonts['normal'])
        self.total_accrued_label.pack(padx=5, pady=5)
        
        frame2 = ct.CTkFrame(info_frame, corner_radius=10)
        frame2.pack(fill='both', padx=10, pady=5)
        ct.CTkLabel(frame2, text='НДФЛ:', font=self.fonts['small']).pack(padx=5, pady=5)
        self.total_ndfl_label = ct.CTkLabel(frame2, text='0 ₽', font=self.fonts['normal'])
        self.total_ndfl_label.pack(padx=5, pady=5)
        
        frame3 = ct.CTkFrame(info_frame, corner_radius=10)
        frame3.pack(fill='both', padx=10, pady=5)
        ct.CTkLabel(frame3, text='Страховые взносы:', font=self.fonts['small']).pack(padx=5, pady=5)
        self.total_insurance_label = ct.CTkLabel(frame3, text='0 ₽', font=self.fonts['normal'])
        self.total_insurance_label.pack(padx=5, pady=5)
        
        frame4 = ct.CTkFrame(info_frame, corner_radius=10)
        frame4.pack(fill='both', padx=10, pady=5)
        ct.CTkLabel(frame4, text='К выплате:', font=self.fonts['small']).pack(padx=5, pady=5)
        self.total_pay_label = ct.CTkLabel(frame4, text='0 ₽', font=self.fonts['normal'])
        self.total_pay_label.pack(padx=5, pady=5)
        
        self.load_calculation_data()

    def on_year_change(self, choice):
        self.load_calculation_data()
    
    def on_month_change(self, choice):
        self.load_calculation_data()

    def load_calculation_data(self):
        """загрузка данных для расчета (только активные сотрудники)"""
        year = int(self.year_var.get())
        month = int(self.month_var.get())
        period = self.db.get_or_create_period(month, year)
        
        employees = self.db.get_all_employees(active_only=True)
        payrolls = self.db.get_payroll_by_period(period['id'])
        existing = {p['employee_id']: p for p in payrolls}
        
        data = [["ID", "ФИО", "Должность", "Тип оплаты", "Отработано часов", 
                 "Премия (₽)", "Доп. начисления (₽)", "Удержания (₽)"]]
        
        for emp in employees:
            pay_data = existing.get(emp['id'], {})
            default_hours = 160 if emp['тип_оплаты'] == 'почасовая' else 0
            data.append([
                emp['id'], emp['ФИО'], emp['должность'],
                "Оклад" if emp['тип_оплаты'] == 'оклад' else "Почасовая",
                pay_data.get('отработано_часов', default_hours) or default_hours,
                pay_data.get('премия', 0) or 0,
                pay_data.get('доп_начисления', 0) or 0,
                pay_data.get('удержек', 0) or 0
            ])
        
        self.calc_sheet.set_sheet_data(data)
        self.calc_sheet.set_all_column_widths()
        self.update_payroll_summary(period['id'])

    def update_payroll_summary(self, period_id):
        summary = self.db.get_summary_by_period(period_id)
        self.total_accrued_label.configure(text=f"{summary['total_accrued']:,.2f} ₽")
        self.total_ndfl_label.configure(text=f"{summary['total_ndfl']:,.2f} ₽")
        self.total_insurance_label.configure(text=f"{summary['total_insurance']:,.2f} ₽")
        self.total_pay_label.configure(text=f"{summary['total_to_pay']:,.2f} ₽")

    def calculate_payroll(self):
        year = int(self.year_var.get())
        month = int(self.month_var.get())
        period = self.db.get_or_create_period(month, year)
        
        sheet_data = self.calc_sheet.get_sheet_data()
        employees_data = []
        for row in sheet_data[1:]:
            if len(row) >= 8 and row[0]:
                employees_data.append({
                    'employee_id': row[0],
                    'отработано_часов': float(row[4]) if row[4] else 0,
                    'премия': float(row[5]) if row[5] else 0,
                    'доп_начисления': float(row[6]) if row[6] else 0,
                    'удержек': float(row[7]) if row[7] else 0
                })
        
        results, summary = self.calc.calculate_period_payroll(period['id'], employees_data)
        self.update_payroll_summary(period['id'])
        self.load_calculation_data()
        messagebox.showinfo("Расчет завершен", f"Зарплата за {month:02d}-{year} рассчитана\nОбщая сумма: {summary['total_to_pay']:,.2f} ₽")

    def on_cell_edited(self, event):
        """сохранение отредактированной ячейки"""
        row = event[0]
        col = event[1]
        value = event[2]
        
        sheet_data = self.calc_sheet.get_sheet_data()
        if row < len(sheet_data):
            employee_id = sheet_data[row][0]
            
            field_map = {
                4: 'отработано_часов',
                5: 'премия',
                6: 'доп_начисления',
                7: 'удержек'
            }
            
            if col in field_map:
                field = field_map[col]
                try:
                    year = int(self.year_var.get())
                    month = int(self.month_var.get())
                    period = self.db.get_period(month, year)
                    if not period:
                        period = self.db.get_or_create_period(month, year)
                    
                    payrolls = self.db.get_payroll_by_period(period['id'])
                    existing = {p['employee_id']: p for p in payrolls}
                    
                    if employee_id in existing:
                        payroll_data = existing[employee_id]
                        new_value = float(value) if value else 0
                        
                        update_data = {
                            'employee_id': employee_id,
                            'period_id': period['id'],
                            'отработано_часов': payroll_data.get('отработано_часов', 0),
                            'премия': payroll_data.get('премия', 0),
                            'доп_начисления': payroll_data.get('доп_начисления', 0),
                            'удержек': payroll_data.get('удержек', 0),
                            'начислено_всего': payroll_data.get('начислено_всего', 0),
                            'НДФЛ': payroll_data.get('НДФЛ', 0),
                            'страховые_взносы': payroll_data.get('страховые_взносы', 0)
                        }
                        update_data[field] = new_value
                        
                        employee = self.db.get_employee(employee_id)
                        if employee:
                            result = self.calc.calculate_employee_salary(
                                employee,
                                worked_hours=update_data['отработано_часов'],
                                bonus=update_data['премия'],
                                additional=update_data['доп_начисления'],
                                deduction=update_data['удержек']
                            )
                            update_data['начислено_всего'] = result['начислено_всего']
                            update_data['НДФЛ'] = result['НДФЛ']
                            update_data['страховые_взносы'] = result['страховые_взносы']
                        
                        self.db.save_payroll(update_data)
                        self.update_payroll_summary(period['id'])
                        self.load_calculation_data()
                        
                except ValueError:
                    messagebox.showerror("Ошибка", "Введите числовое значение")

    # --- вкладка ведомость ---
    def setup_vedomosti_tab(self):
        ved_tab = self.body_tabview.tab("Ведомость")
        
        header_frame = ct.CTkFrame(ved_tab, height=50, corner_radius=0)
        header_frame.pack(fill='x')
        
        ct.CTkLabel(header_frame, text='Ведомость по сотрудникам', font=self.fonts['normal']).pack(anchor='nw', padx=5, pady=5)
        ct.CTkButton(header_frame, text="Экспорт", height=35, corner_radius=15,
                    font=self.fonts['button'], command=self.export_vedomosti).pack(anchor='ne', pady=15)
        
        # выбор года
        year_frame = ct.CTkFrame(header_frame)
        year_frame.pack(anchor='nw', padx=5, pady=5)
        ct.CTkLabel(year_frame, text="Год:", font=self.fonts['small']).pack(side="left", padx=5)
        self.ved_year_var = ct.StringVar(value=str(datetime.now().year))
        self.ved_year_combo = ct.CTkComboBox(year_frame, values=[str(y) for y in range(2023, 2027)],
                                              variable=self.ved_year_var, width=80, command=self.on_ved_year_change)
        self.ved_year_combo.pack(side="left", padx=5)
        
        search_frame = ct.CTkFrame(header_frame, corner_radius=10)
        search_frame.pack(fill='both', padx=10, anchor='w', expand=True)
        
        self.vedomosti_search = ct.CTkEntry(search_frame, placeholder_text='Поиск по фамилии или ИНН', font=self.fonts['small'])
        self.vedomosti_search.pack(padx=5, pady=5, fill='x')
        self.vedomosti_search.bind('<KeyRelease>', self.filter_vedomosti)
        
        self.vedomosti_sheet = tksheet.Sheet(search_frame, data=[])
        self.vedomosti_sheet.pack(fill="both", side='left', expand=True)
        self.vedomosti_sheet.set_options(font=self.fonts['table'])
        self.vedomosti_sheet.enable_bindings("single_select", "copy")
        
        self.load_vedomosti_data()

    def on_ved_year_change(self, choice):
        self.load_vedomosti_data()

    def load_vedomosti_data(self):
        """загрузка ведомости (только активные сотрудники)"""
        year = int(self.ved_year_var.get())
        employees = self.db.get_all_employees(active_only=True)
        months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", 
                  "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        
        self.vedomosti_raw_data = []
        for emp in employees:
            row = [emp['id'], emp['ФИО'], emp['ИНН'] or "-", emp['должность']]
            total_year = 0
            monthly = {}
            
            payrolls = self.db.get_payroll_by_employee(emp['id'])
            for p in payrolls:
                if p['год'] == year:
                    monthly[p['месяц']] = p.get('начислено_всего', 0) - p.get('НДФЛ', 0)
            
            for m in range(1, 13):
                amount = monthly.get(m, 0)
                row.append(f"{amount:,.2f}")
                total_year += amount
            row.append(f"{total_year:,.2f}")
            self.vedomosti_raw_data.append(row)
        
        self.vedomosti_display_data = self.vedomosti_raw_data.copy()
        self.update_vedomosti_display()

    def update_vedomosti_display(self):
        months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", 
                  "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        header = [["ID", "ФИО", "ИНН", "Должность"] + months + ["ИТОГО"]]
        self.vedomosti_sheet.set_sheet_data(header + self.vedomosti_display_data)
        self.vedomosti_sheet.set_all_column_widths()

    def filter_vedomosti(self, event):
        text = self.vedomosti_search.get().lower()
        if not text:
            self.vedomosti_display_data = self.vedomosti_raw_data.copy()
        else:
            self.vedomosti_display_data = [row for row in self.vedomosti_raw_data 
                                          if text in row[1].lower() or text in row[2].lower()]
        self.update_vedomosti_display()

    # --- вкладка отчет ---
    def setup_report_tab(self):
        report_tab = self.body_tabview.tab("Отчет")
        
        header_frame = ct.CTkFrame(report_tab, height=50, corner_radius=0)
        header_frame.pack(fill='x')
        
        ct.CTkLabel(header_frame, text='Сводный отчет', font=self.fonts['normal']).pack(anchor='nw', padx=5, pady=5)
        
        # выбор года
        year_frame = ct.CTkFrame(header_frame)
        year_frame.pack(anchor='nw', padx=5, pady=5)
        ct.CTkLabel(year_frame, text="Год:", font=self.fonts['small']).pack(side="left", padx=5)
        self.report_year_var = ct.StringVar(value=str(datetime.now().year))
        self.report_year_combo = ct.CTkComboBox(year_frame, values=[str(y) for y in range(2023, 2027)],
                                                 variable=self.report_year_var, width=80, command=self.on_report_year_change)
        self.report_year_combo.pack(side="left", padx=5)
        
        self.report_total_label = ct.CTkLabel(header_frame, text='На выплаты (итого): 0 ₽', font=self.fonts['small'])
        self.report_total_label.pack(anchor='nw', padx=5, pady=5)
        self.report_ndfl_label = ct.CTkLabel(header_frame, text='НДФЛ: 0 ₽', font=self.fonts['small'])
        self.report_ndfl_label.pack(anchor='nw', padx=5, pady=5)
        self.report_insurance_label = ct.CTkLabel(header_frame, text='Страховые взносы: 0 ₽', font=self.fonts['small'])
        self.report_insurance_label.pack(anchor='nw', padx=5, pady=5)
        
        ct.CTkButton(header_frame, text="Экспорт", height=35, corner_radius=15,
                    font=self.fonts['button'], command=self.export_report).pack(anchor='se', pady=15)
        
        self.report_sheet = tksheet.Sheet(header_frame, data=[])
        self.report_sheet.pack(fill="both", expand=True)
        self.report_sheet.set_options(font=self.fonts['table'])
        self.report_sheet.enable_bindings("single_select", "copy")
        
        self.load_report_data()

    def on_report_year_change(self, choice):
        self.load_report_data()

    def load_report_data(self):
        year = int(self.report_year_var.get())
        periods = self.db.get_periods_by_year(year)
        data = [["Период", "Сотрудников", "Начислено (₽)", "НДФЛ (₽)", "Страховые взносы (₽)", "К выплате (₽)"]]
        
        for p in periods:
            s = self.db.get_summary_by_period(p['id'])
            data.append([f"{p['месяц']:02d}/{p['год']}", s['count_employees'],
                        f"{s['total_accrued']:,.2f}", f"{s['total_ndfl']:,.2f}",
                        f"{s['total_insurance']:,.2f}", f"{s['total_to_pay']:,.2f}"])
        
        year_summary = self.db.get_summary_by_year(year)
        data.append(["ИТОГО ЗА ГОД", "", f"{year_summary['total_accrued']:,.2f}", 
                    f"{year_summary['total_ndfl']:,.2f}", f"{year_summary['total_insurance']:,.2f}", 
                    f"{year_summary['total_to_pay']:,.2f}"])
        
        self.report_sheet.set_sheet_data(data)
        self.report_sheet.set_all_column_widths()
        
        self.report_total_label.configure(text=f"На выплаты (итого): {year_summary['total_to_pay']:,.2f} ₽")
        self.report_ndfl_label.configure(text=f"НДФЛ: {year_summary['total_ndfl']:,.2f} ₽")
        self.report_insurance_label.configure(text=f"Страховые взносы: {year_summary['total_insurance']:,.2f} ₽")

    # --- экспорт ---
    def export_to_excel(self, data, filename, period_info=""):
        from tkinter import filedialog
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        if period_info:
            filename = f"{filename}_{period_info}"
        
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel файлы", "*.xlsx")], initialfile=filename)
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Данные"
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for i, row in enumerate(data, 1):
                for j, val in enumerate(row, 1):
                    cell = ws.cell(row=i, column=j, value=val)
                    if i == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
            
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
            
            wb.save(path)
            messagebox.showinfo("Успех", f"Сохранено в {path}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def export_employees(self):
        data = self.staff_sheet.get_sheet_data()
        if data and len(data) > 1:
            self.export_to_excel(data, "сотрудники")
    
    def export_calculation(self):
        data = self.calc_sheet.get_sheet_data()
        if data and len(data) > 1:
            period_info = f"{self.month_var.get()}-{self.year_var.get()}"
            self.export_to_excel(data, "расчет_зарплаты", period_info)
    
    def export_vedomosti(self):
        data = self.vedomosti_sheet.get_sheet_data()
        if data and len(data) > 1:
            self.export_to_excel(data, f"ведомость_{self.ved_year_var.get()}")
    
    def export_report(self):
        data = self.report_sheet.get_sheet_data()
        if data and len(data) > 1:
            self.export_to_excel(data, f"отчет_{self.report_year_var.get()}")

    # --- загрузка иконок ---
    def load_icon(self, name_ico, size_im):
        if getattr(sys, 'frozen', False):
            path = os.path.join(os.path.dirname(sys.executable), 'app', name_ico)
        else:
            path = os.path.join(os.path.dirname(os.path.abspath(__file__)), name_ico)
        try:
            return ct.CTkImage(light_image=Image.open(path), size=(size_im, size_im))
        except:
            return None